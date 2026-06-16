"""
generar_datos_prueba.py
-----------------------
Genera el Excel de datos de prueba con el formato de DatosPruebas2026_1.
Llena las columnas de PRUEBAS 3-PARTICIONES > Geometric con los resultados
de KPartitionGeometricSIA para el sistema N10A (estado 1000000000).

Uso (desde Method2_Dynamic_Programming_Reformulation/):
    uv run python generar_datos_prueba.py
"""

import sys
import argparse
import time
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Paths ---
PROJECT_ROOT = Path(__file__).resolve().parents[3]
METHOD2_ROOT = Path(__file__).resolve().parent

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.models.base.application import aplicacion
from src.controllers.manager import Manager
from KParticiones.GeoMIP_K.src.strategies.geometric_k import KPartitionGeometricSIA

SAMPLES_DIR = PROJECT_ROOT / "GeoMIP" / "data" / "samples"
RESULTS_DIR = PROJECT_ROOT / "GeoMIP" / "results"

# --- Configuraciones de sistemas disponibles ---
CONFIGS = {
    "N10A": {
        "estado_inicial": "1000000000",
        "nodos": list("ABCDEFGHIJ"),
        "pruebas": [
            ("ABCDEFGHIJ", "ABCDEFGHIJ"),
            ("ABCDEFGHIJ", "ABCDEFGHI"),
            ("ABCDEFGHIJ", "BCDEFGHIJ"),
            ("ABCDEFGHIJ", "BCDEFGHI"),
            ("ABCDEFGHIJ", "ABDEGHJ"),
            ("ABCDEFGHIJ", "ACEGI"),
            ("ABCDEFGHIJ", "BDFHJ"),
            ("ABCDEFGHI",  "ABCDEFGHIJ"),
            ("ABCDEFGHI",  "ABCDEFGHI"),
            ("ABCDEFGHI",  "BCDEFGHIJ")
        ],
    },
    "N15A": {
        "estado_inicial": "100000000000000",
        "nodos": list("ABCDEFGHIJKLMNO"),
        "pruebas": [
            ("ABCDEFGHIJKLMNO", "ABCDEFGHIJKLMNO"),
            ("ABCDEFGHIJKLMNO", "ABCDEFGHIJKLMN"),
            ("ABCDEFGHIJKLMNO", "BCDEFGHIJKLMNO"),
            ("ABCDEFGHIJKLMNO", "BCDEFGHIJKLMN"),
            ("ABCDEFGHIJKLMNO", "ABDEGHJKMN"),
            ("ABCDEFGHIJKLMNO", "ACEGIKMO"),
            ("ABCDEFGHIJKLMNO", "BDFHJLN"),
            ("ABCDEFGHIJKLMN",  "ABCDEFGHIJKLMNO"),
            ("ABCDEFGHIJKLMN",  "ABCDEFGHIJKLMN")
         #  ("ABCDEFGHIJKLMN",  "BCDEFGHIJKLMNO"),
         #  ("ABCDEFGHIJKLMN",  "BCDEFGHIJKLMN"),
         #   ("ABCDEFGHIJKLMN",  "ABDEGHJKMN"),
         #   ("ABCDEFGHIJKLMN",  "ACEGIKMO"),
         #   ("ABCDEFGHIJKLMN",  "BDFHJLN"),
         #  ("BCDEFGHIJKLMNO",  "ABCDEFGHIJKLMNO"),
        ],
    },
    "N20A": {
        "estado_inicial": "10000000000000000000",
        "nodos": list("ABCDEFGHIJKLMNOPQRST"),
        "pruebas": [
            ("ABCDEFGHIJKLMNOPQRST", "ABCDEFGHIJKLMNOPQRST"),
            ("ABCDEFGHIJKLMNOPQRST", "ABCDEFGHIJKLMNOPQRS"),
            ("ABCDEFGHIJKLMNOPQRST", "BCDEFGHIJKLMNOPQRST"),
            ("ABCDEFGHIJKLMNOPQRST", "BCDEFGHIJKLMNOPQRS"),
            ("ABCDEFGHIJKLMNOPQRST", "ABDEGHIJLMOPQRST"),
            ("ABCDEFGHIJKLMNOPQRST", "ACEGIKMOQRS"),
            ("ABCDEFGHIJKLMNOPQRST", "BDFHJLNPRT"),
            ("ABCDEFGHIJKLMNOPQRS",  "ABCDEFGHIJKLMNOPQRST")
            #("ABCDEFGHIJKLMNOPQRS",  "ABCDEFGHIJKLMNOPQRS"),
            #("ABCDEFGHIJKLMNOPQRS",  "BCDEFGHIJKLMNOPQRST"),
            #("ABCDEFGHIJKLMNOPQRS",  "BCDEFGHIJKLMNOPQRS"),
            #("ABCDEFGHIJKLMNOPQRS",  "ABDEGHIJLMOPQRST"),
            #("ABCDEFGHIJKLMNOPQRS",  "ACEGIKMOQRS"),
            #("ABCDEFGHIJKLMNOPQRS",  "BDFHJLNPRT"),
            #("BCDEFGHIJKLMNOPQRST",  "ABCDEFGHIJKLMNOPQRST"),
        ],
    },
}


def letras_a_bits(letras: str, nodos: list) -> str:
    bits = ["0"] * len(nodos)
    for c in letras.upper():
        if c in nodos:
            bits[nodos.index(c)] = "1"
    return "".join(bits)

TIMEOUT_SEG = 300


def ejecutar_geometric_k(tpm, estado_ini, condicion, alcance_bits, mec_bits, k=3):
    try:
        gestor = Manager(estado_inicial=estado_ini)
        analizador = KPartitionGeometricSIA(gestor, k=k)
        t0 = time.perf_counter()
        sol = analizador.aplicar_estrategia(condicion, alcance_bits, mec_bits, tpm)
        t_total = time.perf_counter() - t0
        return str(sol.particion), round(float(sol.perdida), 6), round(t_total, 4)
    except Exception as e:
        return f"ERROR: {e}", None, None


def construir_excel(filas_resultado: list, sistema: str, nodos: list, estado_ini: str, ruta_salida: Path, k: int = 3):
    sistema_str = "".join(nodos)
    wb = Workbook()
    ws = wb.active
    ws.title = sistema

    # --- Colores ---
    fill_verde  = PatternFill("solid", fgColor="00B050")
    fill_amarillo = PatternFill("solid", fgColor="FFFF00")
    fill_rojo   = PatternFill("solid", fgColor="FF0000")
    fill_azul   = PatternFill("solid", fgColor="00B0F0")
    fill_header = PatternFill("solid", fgColor="D9D9D9")

    negrita = Font(bold=True)
    centro  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def celda(ws, fila, col, valor, fill=None, bold=False, alineacion=None):
        c = ws.cell(row=fila, column=col, value=valor)
        if fill:
            c.fill = fill
        if bold:
            c.font = Font(bold=True)
        c.alignment = alineacion or centro
        return c

    # --- Fila 1: estado inicial ---
    ws.cell(row=1, column=1, value="").alignment = centro
    ws.cell(row=1, column=2, value=estado_ini).alignment = centro

    # --- Fila 2: Sistema ---
    ws.cell(row=2, column=1, value="Sistema:").font = negrita
    ws.cell(row=2, column=2, value=sistema_str)

    # --- Fila 3: Sistema Candidato ---
    ws.cell(row=3, column=1, value="Sistema Candidato:").font = negrita
    ws.cell(row=3, column=2, value=sistema_str)

    # --- Fila 4: grupos principales ---
    # Col D-F: PRUEBAS BIPARTICIONES (verde)
    ws.merge_cells("D4:I4")
    celda(ws, 4, 4, "PRUEBAS BIPARTICIONES", fill=fill_verde, bold=True)

    # Col J-O: PRUEBAS 3-PARTICIONES (amarillo)
    ws.merge_cells("J4:O4")
    celda(ws, 4, 10, f"PRUEBAS {k}-PARTICIONES", fill=fill_amarillo, bold=True)

    # --- Fila 5: estrategias ---
    ws.merge_cells("D5:F5")
    celda(ws, 5, 4, "QNodes", fill=fill_rojo, bold=True)
    ws.merge_cells("G5:I5")
    celda(ws, 5, 7, "Geometric", fill=fill_azul, bold=True)
    ws.merge_cells("J5:L5")
    celda(ws, 5, 10, "QNodes", fill=fill_rojo, bold=True)
    ws.merge_cells("M5:O5")
    celda(ws, 5, 13, "Geometric", fill=fill_azul, bold=True)

    # --- Fila 6: columnas detalle ---
    headers = [
        (1, "#Prueba"), (2, "Alcance o Purview"), (3, "Mecanismo(t)"),
        (4, "Partición"), (5, "Pérdida"), (6, "Tiempo"),
        (7, "Partición"), (8, "Pérdida"), (9, "Tiempo"),
        (10, "Partición"), (11, "Pérdida"), (12, "Tiempo"),
        (13, "Partición"), (14, "Pérdida"), (15, "Tiempo"),
    ]
    for col, titulo in headers:
        celda(ws, 6, col, titulo, fill=fill_header, bold=True)

    # --- Filas de datos ---
    for idx, (prueba_num, alcance_l, mec_l, part_k3, perdida_k3, tiempo_k3) in enumerate(filas_resultado):
        fila = 7 + idx
        ws.cell(row=fila, column=1,  value=prueba_num).alignment = centro
        ws.cell(row=fila, column=2,  value=alcance_l).alignment = centro
        ws.cell(row=fila, column=3,  value=mec_l).alignment = centro
        # Cols 4-9: biparticiones (dejar vacías — no son de este script)
        # Cols 13-15: Geometric k=3
        c_part = ws.cell(row=fila, column=13, value=part_k3)
        c_part.alignment = Alignment(horizontal="left", wrap_text=True)
        ws.cell(row=fila, column=14, value=perdida_k3).alignment = centro
        ws.cell(row=fila, column=15, value=tiempo_k3).alignment = centro

    # Anchos de columna
    anchos = {1: 8, 2: 18, 3: 18, 4: 35, 5: 10, 6: 14,
              7: 35, 8: 10, 9: 14, 10: 35, 11: 10, 12: 14,
              13: 40, 14: 10, 15: 14}
    for col, ancho in anchos.items():
        ws.column_dimensions[get_column_letter(col)].width = ancho

    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta_salida)
    print(f"\n  Excel guardado en: {ruta_salida}")


def main():
    parser = argparse.ArgumentParser(description="Generar datos de prueba GeoMIP k-particiones")
    parser.add_argument(
        "--sistema", type=str, choices=list(CONFIGS.keys()), default="N10A",
        help="Sistema a procesar: N10A, N15A o N20A (default: N10A)"
    )
    parser.add_argument(
        "--k", type=int, choices=[3, 4, 5], default=3,
        help="Numero de particiones: 3, 4 o 5 (default: 3)"
    )
    args = parser.parse_args()

    cfg = CONFIGS[args.sistema]
    estado_ini  = cfg["estado_inicial"]
    nodos       = cfg["nodos"]
    pruebas     = cfg["pruebas"]
    k           = args.k

    aplicacion.profiler_habilitado = False

    tpm_path = SAMPLES_DIR / f"{args.sistema}.csv"
    if not tpm_path.exists():
        print(f"No se encontro: {tpm_path}")
        return

    tpm = np.genfromtxt(tpm_path, delimiter=",")
    condicion = "1" * len(estado_ini)

    print(f"\n{'='*55}")
    print(f"  Generando datos de prueba — {args.sistema}  k={k}")
    print(f"{'='*55}")

    filas = []
    for i, (alcance_l, mec_l) in enumerate(pruebas, start=1):
        alcance_bits = letras_a_bits(alcance_l, nodos)
        mec_bits     = letras_a_bits(mec_l, nodos)

        print(f"  Prueba {i:02d}: alcance={alcance_l}  mec={mec_l} ... ", end="", flush=True)
        part, perdida, tiempo = ejecutar_geometric_k(
            tpm, estado_ini, condicion, alcance_bits, mec_bits, k=k
        )
        estado = f"perdida={perdida}  t={tiempo}s" if perdida is not None else part
        print(estado)
        filas.append((i, alcance_l, mec_l, part, perdida, tiempo))

    salida = RESULTS_DIR / f"DatosPruebas_Geometric_k{k}_{args.sistema}.xlsx"
    construir_excel(filas, args.sistema, nodos, estado_ini, salida, k=k)
    print(f"{'='*55}\n")


if __name__ == "__main__":
    main()
